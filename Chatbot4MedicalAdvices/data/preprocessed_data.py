import openpyxl

workbook = openpyxl.load_workbook('datasets.xlsx')
sheet1 = workbook.active

nrows = sheet1.max_row
ncols = sheet1.max_column
print(nrows,ncols)

for i in range(2, nrows + 1):
    res = "F" + str(i)
    # print(res)
    # print(sheet1[res].value)
    cellv = sheet1[res].value
    if cellv == "暂无":
        sheet1[res].value = "[]"
    else:
        list = cellv.split(',')
        list = str(list).replace("'", "\"")
        # print(list)
        sheet1[res].value = list

for i in range(2, nrows + 1):
    res = "J" + str(i)
    # print(res)
    # print(sheet1[res].value)
    cellv = sheet1[res].value
    if cellv == "暂无":
        sheet1[res].value = "[]"
    else:
        list = cellv.split(',')
        list = str(list).replace("'", "\"")
        # print(list)
        sheet1[res].value = list

for i in range(2, nrows + 1):
    res = "K" + str(i)
    # print(res)
    # print(sheet1[res].value)
    cellv = sheet1[res].value
    if cellv == "暂无":
        sheet1[res].value = "[]"
    else:
        list = cellv.split(',')
        list = str(list).replace("'", "\"")
        # print(list)
        sheet1[res].value = list

for i in range(2, nrows + 1):
    res = "C" + str(i)
    sheet1[res].value = "[]"

for i in range(2, nrows + 1):
    res = "L" + str(i)
    sheet1[res].value = "[]"

for i in range(2, nrows + 1):
    res = "P" + str(i)
    sheet1[res].value = "[]"

for i in range(2, nrows + 1):
    res = "Q" + str(i)
    sheet1[res].value = "[]"

for i in range(2, nrows + 1):
    res = "R" + str(i)
    sheet1[res].value = "[]"

workbook.save(filename = "preprocessed_datasets.xlsx")


f = open("final_data2.txt","r",encoding='utf-8')
text = f.read()
# print(text)
text_processed = text.replace('\"[','[')
text_processed = text_processed.replace(']\"',']')
text_processed = text_processed.replace('\\\"', '\"')
text_processed = text_processed.replace('\\\\\"','\"')
text_processed = text_processed.replace('\"{','{')
text_processed = text_processed.replace('}\"', '}')
fw = open("data_processed.txt", 'w',encoding='utf-8')
fw.write(text_processed)
print(text_processed)